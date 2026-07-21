"""Render the optimized timetable CSV as university-ready Excel and PDF grids."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Table, TableStyle


ROOT = Path(__file__).resolve().parents[1]
RESULTS = ROOT / "results"
SOURCE = RESULTS / "optimized_timetable.csv"
SECTIONS = ["CSE-A", "CSE-B", "CSE-C"]
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
PERIODS = [f"P{number}" for number in range(1, 9)]


def load_grids() -> dict[str, list[list[str]]]:
    """Create display-only grids from the existing generated timetable rows."""
    if not SOURCE.exists():
        raise FileNotFoundError(f"Timetable CSV not found: {SOURCE}")
    data = pd.read_csv(SOURCE).fillna("")
    required = {"Section", "Course", "Faculty", "Room", "Day", "Period"}
    missing = required.difference(data.columns)
    if missing:
        raise ValueError(f"CSV is missing required columns: {', '.join(sorted(missing))}")

    grids = {}
    for section in SECTIONS:
        grid = [["FREE" for _ in PERIODS] for _ in DAYS]
        rows = data[data["Section"].astype(str) == section]
        for _, item in rows.iterrows():
            day, period = str(item["Day"]), str(item["Period"])
            if day not in DAYS or period not in PERIODS:
                continue
            value = f"{item['Course']}\n{item['Faculty']}\n{item['Room']}"
            row_index, column_index = DAYS.index(day), PERIODS.index(period)
            grid[row_index][column_index] = (
                value if grid[row_index][column_index] == "FREE"
                else f"{grid[row_index][column_index]}\n---\n{value}"
            )
        grids[section] = grid
    return grids


def create_excel(grids: dict[str, list[list[str]]]) -> Path:
    """Write one styled worksheet per requested section."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    navy = "17365D"
    blue = "D9EAF7"
    free = "F2F2F2"
    thin = Side(style="thin", color="7F8C8D")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for section, grid in grids.items():
        sheet = workbook.create_sheet(section)
        sheet.merge_cells("A1:I1")
        title = sheet["A1"]
        title.value = f"{section} Weekly Timetable"
        title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill("solid", fgColor=navy)
        title.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28
        headers = ["Day", *PERIODS]
        for column, label in enumerate(headers, 1):
            cell = sheet.cell(3, column, label)
            cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        for row, (day, values) in enumerate(zip(DAYS, grid), 4):
            day_cell = sheet.cell(row, 1, day)
            day_cell.font = Font(name="Calibri", bold=True)
            day_cell.fill = PatternFill("solid", fgColor=blue)
            day_cell.alignment = Alignment(horizontal="center", vertical="center")
            day_cell.border = border
            for column, value in enumerate(values, 2):
                cell = sheet.cell(row, column, value)
                cell.font = Font(name="Calibri", size=10, bold=value == "FREE")
                cell.fill = PatternFill("solid", fgColor=free if value == "FREE" else "FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            sheet.row_dimensions[row].height = 60
        sheet.column_dimensions["A"].width = 15
        for column in range(2, 10):
            sheet.column_dimensions[get_column_letter(column)].width = 20
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "B4"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
    target = RESULTS / "sample_timetable.xlsx"
    workbook.save(target)
    return target


def create_pdf(grids: dict[str, list[list[str]]]) -> Path:
    """Write a clean landscape PDF with one section timetable per page."""
    target = RESULTS / "sample_timetable.pdf"
    document = SimpleDocTemplate(
        str(target), pagesize=landscape(A4),
        leftMargin=1.0 * cm, rightMargin=1.0 * cm,
        topMargin=1.0 * cm, bottomMargin=1.0 * cm,
    )
    styles = getSampleStyleSheet()
    story = []
    for number, (section, grid) in enumerate(grids.items()):
        story.append(Paragraph(f"{section} Weekly Timetable", styles["Title"]))
        matrix = [["Day", *PERIODS]] + [[day, *values] for day, values in zip(DAYS, grid)]
        table = Table(matrix, colWidths=[2.4 * cm] + [3.1 * cm] * 8, rowHeights=[0.8 * cm] + [2.6 * cm] * 5)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#D9EAF7")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#7F8C8D")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("LEADING", (1, 1), (-1, -1), 10),
        ]))
        story.append(table)
        if number < len(grids) - 1:
            story.append(PageBreak())
    document.build(story)
    return target


def main() -> None:
    grids = load_grids()
    create_excel(grids)
    create_pdf(grids)


if __name__ == "__main__":
    main()
