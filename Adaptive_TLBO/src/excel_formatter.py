"""Professional Excel workbook generation for university timetables."""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """Creates a polished university-style multi-sheet timetable workbook."""

    TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FILL = PatternFill("solid", fgColor="17365D")
    SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
    ALT_FILL = PatternFill("solid", fgColor="F4F8FB")
    FREE_FILL = PatternFill("solid", fgColor="E7E6E6")
    LUNCH_FILL = PatternFill("solid", fgColor="FCE4D6")
    LAB_FILL = PatternFill("solid", fgColor="E2F0D9")
    THEORY_FILL = PatternFill("solid", fgColor="DDEBF7")
    WHITE_BOLD = Font(name="Calibri", color="FFFFFF", bold=True)
    TITLE_FONT = Font(name="Calibri", size=16, color="FFFFFF", bold=True)
    NORMAL_FONT = Font(name="Calibri", size=11)
    BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="17365D")
    BORDER = Border(
        left=Side(style="thin", color="9EADBC"),
        right=Side(style="thin", color="9EADBC"),
        top=Side(style="thin", color="9EADBC"),
        bottom=Side(style="thin", color="9EADBC"),
    )

    def __init__(self, dataset):
        self.dataset = dataset
        self.days = dataset.get_days()
        self.periods = dataset.get_periods()

    def save(
        self,
        timetable,
        summary: Dict[str, object],
        statistics: Dict[str, object],
        file_path,
    ) -> None:
        file_path = Path(file_path)
        file_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        wb.remove(wb.active)
        # Required print-facing tabs are deliberately created first.
        for section in self.dataset.get_sections():
            self._section_timetable_sheet(wb, section, timetable)
        self._faculty_timetable_sheet(wb, timetable)

 
        self._room_timetable_sheet(wb, timetable)

 
        wb["Room Timetable"].title = "Room Utilization"

        self._statistics_sheet(wb, statistics)
        wb["Statistics"].title = "Faculty Workload Summary"
        self._summary_sheet(wb, summary)
        self._master_data_sheet(wb, timetable)
        for sheet in wb.worksheets:
            sheet.page_setup.orientation = "landscape"
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.sheet_view.showGridLines = False
        wb.save(file_path)

    def _summary_sheet(self, wb, summary: Dict[str, object]) -> None:
        ws = wb.create_sheet("Summary")
        self._title(ws, "Adaptive TLBO University Timetable Summary", 2)
        ws.append([])
        for key, value in summary.items():
            ws.append([key, value])
        self._style_key_value(ws, start_row=3, end_row=ws.max_row)
        ws.freeze_panes = "A3"

    def _master_data_sheet(self, wb, timetable) -> None:
        ws = wb.create_sheet("Master Data")
        self._title(ws, "Master Data", 11)
        headers = [
            "Section",
            "Course",
            "Faculty",
            "Room",
            "Day",
            "Time",
            "Period",
            "Hours",
            "Priority",
            "SubjectID",
            "Type",
        ]
        ws.append(headers)
        for entry in self._sorted_timetable(timetable):
            ws.append([
                entry["Section"],
                entry["Course"],
                entry["Faculty"],
                entry["Room"],
                entry["Day"],
                int(entry["Time"]),
                f"P{int(entry['Time'])}",
                entry["Hours"],
                entry["Priority"],
                entry["SubjectID"],
                entry.get("Type", "Theory"),
            ])
        self._style_grid(ws, header_row=2, freeze="A3", filters=True)

    def _section_timetable_sheet(self, wb, section: str, timetable) -> None:
        ws = wb.create_sheet(f"{section} Timetable")
        columns = self._section_columns()
        self._title(ws, f"{section} Timetable", len(columns))
        ws.append(columns)

        entries = {
            (entry["Day"], int(entry["Time"])): entry
            for entry in timetable
            if entry["Section"] == section
        }

        for day in self.days:
            row = [day]
            for column in columns[1:]:
                if column == "Lunch Break":
                    row.append("LUNCH BREAK")
                    continue
                period = int(column[1:])
                entry = entries.get((day, period))
                if entry:
                    row.append(
                        f"{entry['Course']}\n"
                        f"{entry['Faculty']}\n"
                        f"{entry['Room']}"
                    )
                else:
                    row.append("FREE")
            ws.append(row)

        self._style_timetable(ws, title_columns=len(columns), merge_lunch=True)
        self._highlight_section_cells(ws, timetable, section)

    def _faculty_timetable_sheet(self, wb, timetable) -> None:
        ws = wb.create_sheet("Faculty Timetable")
        columns = ["Faculty"] + self._slot_headers()
        self._title(ws, "Faculty Timetable", len(columns))
        ws.append(columns)

        faculty_names = sorted(self.dataset.get_faculty()["FacultyName"].astype(str).unique())
        for faculty in faculty_names:
            allocation = self._faculty_allocation(timetable, faculty)
            ws.append([faculty] + [allocation.get(slot, "FREE") for slot in self._slot_headers()])

        self._style_wide_timetable(ws, len(columns))

    def _room_timetable_sheet(self, wb, timetable) -> None:
        ws = wb.create_sheet("Room Timetable")
        columns = ["Room"] + self._slot_headers()
        self._title(ws, "Room Timetable", len(columns))
        ws.append(columns)

        for room in sorted(self.dataset.get_rooms()["RoomID"].astype(str).tolist()):
            allocation = self._room_allocation(timetable, room)
            ws.append([room] + [allocation.get(slot, "FREE") for slot in self._slot_headers()])

        self._style_wide_timetable(ws, len(columns))

    def _statistics_sheet(self, wb, statistics: Dict[str, object]) -> None:
        ws = wb.create_sheet("Statistics")
        self._title(ws, "Timetable Statistics", 2)
        ws.append(["Metric", "Value"])
        for key, value in statistics.items():
            ws.append([key, value])
        self._style_grid(ws, header_row=2, freeze="A3", filters=True)
        if ws.max_row >= 3:
            ws.conditional_formatting.add(
                f"B3:B{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                ),
            )

    def _faculty_allocation(self, timetable, faculty: str) -> Dict[str, str]:
        allocation = {}
        for entry in timetable:
            if entry["Faculty"] != faculty:
                continue
            key = f"{entry['Day']} P{int(entry['Time'])}"
            allocation[key] = (
                f"{entry['Course']}\n"
                f"{entry['Section']}\n"
                f"{entry['Room']}"
            )
        return allocation

    def _room_allocation(self, timetable, room: str) -> Dict[str, str]:
        allocation = {}
        for entry in timetable:
            if str(entry["Room"]) != str(room):
                continue
            key = f"{entry['Day']} P{int(entry['Time'])}"
            allocation[key] = (
                f"{entry['Section']}\n"
                f"{entry['Course']}\n"
                f"{entry['Faculty']}"
            )
        return allocation

    def _style_timetable(self, ws, title_columns: int, merge_lunch: bool) -> None:
        self._style_grid(ws, header_row=2, freeze="B3", filters=False)
        lunch_col = self._find_header_column(ws, "Lunch Break", 2)

        for row in range(3, ws.max_row + 1):
            ws.row_dimensions[row].height = 62
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                value = str(cell.value or "")
                if value == "FREE":
                    cell.fill = self.FREE_FILL
                    cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
                elif value == "LUNCH BREAK":
                    cell.fill = self.LUNCH_FILL
                    cell.font = self.BOLD_FONT
                elif "\n" in value:
                    cell.fill = self.THEORY_FILL

        if lunch_col:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=lunch_col).fill = self.LUNCH_FILL
                ws.cell(row=row, column=lunch_col).font = self.BOLD_FONT
            ws.column_dimensions[get_column_letter(lunch_col)].width = 16
            if merge_lunch and ws.max_row > 3:
                ws.merge_cells(
                    start_row=3,
                    start_column=lunch_col,
                    end_row=ws.max_row,
                    end_column=lunch_col,
                )
                merged = ws.cell(row=3, column=lunch_col)
                merged.value = "LUNCH BREAK"
                merged.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    text_rotation=90,
                    wrap_text=True,
                )
                merged.fill = self.LUNCH_FILL
                merged.font = self.BOLD_FONT

        for column in range(2, title_columns + 1):
            ws.column_dimensions[get_column_letter(column)].width = 22

    def _highlight_section_cells(self, ws, timetable, section: str) -> None:
        header_map = {
            str(ws.cell(row=2, column=column).value): column
            for column in range(1, ws.max_column + 1)
        }
        day_row = {
            str(ws.cell(row=row, column=1).value): row
            for row in range(3, ws.max_row + 1)
        }
        for entry in timetable:
            if entry["Section"] != section:
                continue
            row = day_row.get(entry["Day"])
            column = header_map.get(f"P{int(entry['Time'])}")
            if not row or not column:
                continue
            cell = ws.cell(row=row, column=column)
            if str(entry.get("Type", "Theory")).lower() == "lab":
                cell.fill = self.LAB_FILL
            else:
                cell.fill = self.THEORY_FILL

    def _style_wide_timetable(self, ws, title_columns: int) -> None:
        self._style_grid(ws, header_row=2, freeze="B3", filters=True)
        for row in range(3, ws.max_row + 1):
            ws.row_dimensions[row].height = 52
            for column in range(2, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                if cell.value == "FREE":
                    cell.fill = self.FREE_FILL
                    cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
                else:
                    cell.fill = self.THEORY_FILL
        for column in range(2, title_columns + 1):
            ws.column_dimensions[get_column_letter(column)].width = 18

    def _style_grid(self, ws, header_row: int, freeze: str, filters: bool) -> None:
        for row in range(header_row, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=column)
                cell.border = self.BORDER
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                if row > header_row and row % 2 == 1 and not cell.fill.fill_type:
                    cell.fill = self.ALT_FILL

        for cell in ws[header_row]:
            cell.fill = self.HEADER_FILL
            cell.font = self.WHITE_BOLD

        for column in range(1, ws.max_column + 1):
            letter = get_column_letter(column)
            max_length = max(
                len(str(ws.cell(row=row, column=column).value or "").split("\n")[0])
                for row in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[letter].width = min(max(max_length + 4, 12), 32)

        ws.freeze_panes = freeze
        if filters:
            ws.auto_filter.ref = ws.dimensions

    def _style_key_value(self, ws, start_row: int, end_row: int) -> None:
        for row in range(start_row, end_row + 1):
            for column in range(1, 3):
                cell = ws.cell(row=row, column=column)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if column == 1:
                    cell.font = self.BOLD_FONT
                    cell.fill = self.SUBHEADER_FILL
                else:
                    cell.font = self.NORMAL_FONT
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 28

    def _title(self, ws, title: str, last_column: int) -> None:
        ws.append([title])
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=max(1, last_column),
        )
        cell = ws["A1"]
        cell.fill = self.TITLE_FILL
        cell.font = self.TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = self.BORDER
        ws.row_dimensions[1].height = 28

    def _section_columns(self) -> List[str]:
        columns = ["Day"]
        for period in self.periods:
            if period == 4:
                columns.append("Lunch Break")
            columns.append(f"P{period}")
        return columns

    def _slot_headers(self) -> List[str]:
        return [
            f"{day} P{period}"
            for day in self.days
            for period in self.periods
        ]

    def _sorted_timetable(self, timetable):
        day_order = {day: index for index, day in enumerate(self.days)}
        return sorted(
            timetable,
            key=lambda entry: (
                entry["Section"],
                day_order.get(entry["Day"], 99),
                int(entry["Time"]),
                entry["Course"],
            ),
        )

    @staticmethod
    def _find_header_column(ws, header: str, header_row: int) -> int | None:
        for column in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=column).value == header:
                return column
        return None


def format_excel(file_path):
    """Compatibility wrapper retained for older imports."""
    return file_path

